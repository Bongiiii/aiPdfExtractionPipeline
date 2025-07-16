# -*- coding: utf-8 -*-
"""
Created on Wed Jul  2 16:54:26 2025

@author: bongi
"""

import os
import io
import json
import base64
import time
from pathlib import Path
from typing import List, Dict, Any
import pandas as pd
from PIL import Image
import fitz  # PyMuPDF
from openai import OpenAI
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

# Load API key from .env file
load_dotenv()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

if not OPENAI_API_KEY:
    raise ValueError("OPENAI_API_KEY not found in .env file. Please add OPENAI_API_KEY=your_key_here to your .env file")

class PDFTableExtractor:
    def __init__(self, api_key: str):
        self.client = OpenAI(api_key=api_key)
        self.output_dir = "extracted_tables"
        os.makedirs(self.output_dir, exist_ok=True)
        print(f"Output directory: {os.path.abspath(self.output_dir)}")

    def pdf_to_images(self, pdf_path: str, dpi: int = 300) -> List[Image.Image]:
        doc = fitz.open(pdf_path)
        images = []
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            mat = fitz.Matrix(dpi/72, dpi/72)
            pix = page.get_pixmap(matrix=mat)
            img_data = pix.tobytes("png")
            img = Image.open(io.BytesIO(img_data))
            images.append(img)
        doc.close()
        return images

    def encode_image(self, image: Image.Image) -> str:
        buffer = io.BytesIO()
        image.save(buffer, format='PNG')
        image_bytes = buffer.getvalue()
        return base64.b64encode(image_bytes).decode('utf-8')

    def extract_tables_from_image(self, image: Image.Image, page_num: int) -> List[Dict[str, Any]]:
        base64_image = self.encode_image(image)

        prompt = """
This is a scanned government document page that contains a table without visible borders. Extract the table if and only if it includes species listings in this 3-column format:

- Common Name
- Scientific Name
- Where Found

Your output should only include actual rows in this 3-column format.

Return as:
{
  "tables": [
    {
      "table_name": "species_table",
      "headers": ["Common Name", "Scientific Name", "Where Found"],
      "data": [
        ["Spider monkey", "Ateles spp.", "Central and South America"],
        ...
      ]
    }
  ]
}

Clean spacing, keep scientific names intact, and do not include unrelated narrative text or metadata.
"""

        try:
            response = self.client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/png;base64,{base64_image}"
                                }
                            }
                        ]
                    }
                ],
                max_tokens=4000,
                temperature=0.1
            )

            response_text = response.choices[0].message.content
            print(f"Raw response from OpenAI (Page {page_num + 1}): {response_text[:200]}...")

            if "```json" in response_text:
                json_start = response_text.find("```json") + 7
                json_end = response_text.find("```", json_start)
                response_text = response_text[json_start:json_end].strip()
            elif "```" in response_text:
                json_start = response_text.find("```") + 3
                json_end = response_text.find("```", json_start)
                response_text = response_text[json_start:json_end].strip()

            try:
                result = json.loads(response_text)
                tables = result.get("tables", [])
                for table in tables:
                    table["page_number"] = page_num + 1
                print(f"Successfully extracted {len(tables)} tables from page {page_num + 1}")
                return tables
            except json.JSONDecodeError as e:
                print(f"JSON decode error on page {page_num + 1}: {e}")
                print(f"Response text: {response_text[:500]}...")
                return []

        except Exception as e:
            print(f"Error processing page {page_num + 1}: {e}")
            return []

    def create_excel_file(self, tables: List[Dict[str, Any]], output_path: str):
        if not tables:
            print("No tables to write to Excel file")
            return

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            sheet_names = set()
            for i, table in enumerate(tables):
                base_name = table.get("table_name", f"Table_{i+1}")
                sheet_name = "".join(c for c in base_name if c.isalnum() or c in (' ', '_', '-'))[:31]
                original_name = sheet_name
                counter = 1
                while sheet_name in sheet_names:
                    sheet_name = f"{original_name}_{counter}"[:31]
                    counter += 1
                sheet_names.add(sheet_name)

                headers = table.get("headers", [])
                data = table.get("data", [])

                if not headers and data:
                    headers = [f"Column_{j+1}" for j in range(len(data[0]) if data else 0)]

                df = pd.DataFrame(data, columns=headers) if data else pd.DataFrame(columns=headers)

                df.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]

                # Format headers
                for col in range(1, len(headers) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

        print(f"Excel file created: {output_path}")

    def process_pdf(self, pdf_path: str) -> str:
        pdf_name = Path(pdf_path).stem
        output_excel_path = os.path.join(self.output_dir, f"{pdf_name}_extracted_tables.xlsx")

        print(f"Processing PDF: {pdf_path}")

        try:
            print("Converting PDF to images...")
            images = self.pdf_to_images(pdf_path)
            print(f"Successfully converted PDF to {len(images)} images")

            all_tables = []
            relevant_pages = [4,6]  # Page 7 and 8 in the document (0-indexed)
            for page_num in relevant_pages:
                if page_num >= len(images):
                    continue
                print(f"Processing page {page_num + 1}/{len(images)}...")
                tables = self.extract_tables_from_image(images[page_num], page_num)
                all_tables.extend(tables)
                time.sleep(1)

            if all_tables:
                print(f"Found {len(all_tables)} tables total")
                self.create_excel_file(all_tables, output_excel_path)
            else:
                print("No tables found in the document")
                pd.DataFrame({"Note": ["No tables found in this document"]}).to_excel(
                    output_excel_path, index=False
                )

            return output_excel_path

        except Exception as e:
            print(f"Error processing PDF {pdf_path}: {e}")
            return None

    def process_directory(self, input_dir: str):
        if not os.path.exists(input_dir):
            print(f"Input directory {input_dir} does not exist. Creating it...")
            os.makedirs(input_dir, exist_ok=True)
            print(f"Please place your PDF files in: {os.path.abspath(input_dir)}")
            return

        pdf_files = list(Path(input_dir).glob("*.pdf"))

        if not pdf_files:
            print(f"No PDF files found in {input_dir}")
            print(f"Please place your PDF files in: {os.path.abspath(input_dir)}")
            return

        print(f"Found {len(pdf_files)} PDF files to process")
        successful = 0
        failed = 0

        for pdf_file in pdf_files:
            try:
                result = self.process_pdf(str(pdf_file))
                if result:
                    print(f" Successfully processed: {pdf_file.name}")
                    successful += 1
                else:
                    print(f" Failed to process: {pdf_file.name}")
                    failed += 1
            except Exception as e:
                print(f" Error processing {pdf_file.name}: {e}")
                failed += 1

        print(f"\n Processing complete!")
        print(f" Successfully processed: {successful} files")
        print(f" Failed: {failed} files")
        print(f" Check the '{self.output_dir}' folder for results: {os.path.abspath(self.output_dir)}")


def main():
    INPUT_DIRECTORY = "input_pdfs"
    if not OPENAI_API_KEY:
        print(" Error: OpenAI API key not found!")
        print("Please create a .env file in the same directory as this script with:")
        print("OPENAI_API_KEY=your_actual_api_key_here")
        return

    os.makedirs(INPUT_DIRECTORY, exist_ok=True)
    print(" Starting PDF Table Extractor")
    print(f" Input directory: {os.path.abspath(INPUT_DIRECTORY)}")

    try:
        extractor = PDFTableExtractor(OPENAI_API_KEY)
        extractor.process_directory(INPUT_DIRECTORY)
    except Exception as e:
        print(f" Error initializing extractor: {e}")

if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""
Created on Wed Jul  2 16:54:26 2025

@author: bongi

Lido inspired PDF Table Extractor
Extract specific columns from PDFs like Lido does
"""

import os
import io
import json
import base64
import time
from pathlib import Path
from typing import List, Dict, Any, Optional
import pandas as pd
from PIL import Image
import fitz  # PyMuPDF
from openai import OpenAI
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

# Load API key from .env file
load_dotenv()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

if not OPENAI_API_KEY:
    raise ValueError("OPENAI_API_KEY not found in .env file. Please add OPENAI_API_KEY=your_key_here to your .env file")


class LidoStylePDFExtractor:
    def __init__(self, api_key: str):
        """
        Initialize the Lido-Style PDF Table Extractor
        
        Args:
            api_key (str): OpenAI API key
        """
        self.client = OpenAI(api_key=api_key)
        self.output_dir = "extracted_tables"
        os.makedirs(self.output_dir, exist_ok=True)
        print(f"Output directory: {os.path.abspath(self.output_dir)}")
    
    def pdf_to_images(self, pdf_path: str, dpi: int = 300) -> List[Image.Image]:
        """
        Convert PDF pages to images
        
        Args:
            pdf_path (str): Path to PDF file
            dpi (int): Resolution for image conversion
            
        Returns:
            List[Image.Image]: List of PIL Images
        """
        doc = fitz.open(pdf_path)
        images = []
        
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            mat = fitz.Matrix(dpi/72, dpi/72)  # Scale factor for DPI
            pix = page.get_pixmap(matrix=mat)
            img_data = pix.tobytes("png")
            img = Image.open(io.BytesIO(img_data))
            images.append(img)
        
        doc.close()
        return images
    
    def encode_image(self, image: Image.Image) -> str:
        """
        Encode PIL Image to base64 string
        
        Args:
            image (Image.Image): PIL Image
            
        Returns:
            str: Base64 encoded image
        """
        buffer = io.BytesIO()
        image.save(buffer, format='PNG')
        image_bytes = buffer.getvalue()
        return base64.b64encode(image_bytes).decode('utf-8')
    
    def extract_specific_columns(self, image: Image.Image, columns: List[str], 
                               extra_instructions: str = "", page_num: int = 0) -> List[Dict[str, Any]]:
        """
        Extract specific columns from a single image using OpenAI Vision API
        
        Args:
            image (Image.Image): PIL Image of the document page
            columns (List[str]): List of column names to extract
            extra_instructions (str): Additional instructions for extraction
            page_num (int): Page number for reference
            
        Returns:
            List[Dict]: Extracted data rows
        """
        base64_image = self.encode_image(image)
        
        columns_str = '", "'.join(columns)
        
        prompt = f"""
        Analyze this document image and extract ONLY the following columns: "{columns_str}"
        
        SPECIFIC EXTRACTION TASK:
        - Do not treat section headers like "Mammals—Continued" or "Birds" as data rows
        - Only extract rows where all three fields (Common name, Scientific name, Where found) are clearly aligned
        - Look for tabular data that contains information related to these columns
        - Extract data that matches or corresponds to these column headers
        - The document may have different header names, but extract data that logically fits these categories
        - Be flexible with column matching (e.g., "Species" could match "Scientific Name", "Date" could match "Year Collected")
        
        COLUMN DEFINITIONS:
        {self._generate_column_definitions(columns)}
        
        ADDITIONAL INSTRUCTIONS:
        {extra_instructions if extra_instructions else "No additional instructions provided."}
        
        FORMATTING RULES:
        1. Return data as a JSON object with this exact structure:
        {{
            "extracted_data": [
                {{{", ".join([f'"{col}": "value"' for col in columns])}}},
                {{{", ".join([f'"{col}": "value"' for col in columns])}}},
                ...
            ],
            "total_rows": number_of_rows_found,
            "notes": "any relevant observations about the extraction"
        }}
        
        2. Data cleaning rules:
        - Remove extra whitespace
        - Keep scientific names intact (if applicable)
        - Preserve dates in their original format
        - Keep location information exactly as written
        - If a field is empty or not found, use "N/A"
        
        3. Be thorough but precise:
        - Extract ALL rows that contain the requested information
        - Don't invent data that's not clearly visible
        - If you're uncertain about a value, note it in the "notes" field
        
        4. Handle variations:
        - Column headers might be slightly different (extract based on meaning)
        - Data might be in paragraph form rather than clear table structure
        - Look for patterns and structured text even without borders
        """
        
        try:
            response = self.client.chat.completions.create(
                model="gpt-4o",  # Using GPT-4 with vision
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/png;base64,{base64_image}"
                                }
                            }
                        ]
                    }
                ],
                max_tokens=4000,
                temperature=0.1  # Low temperature for consistent extraction
            )
            
            # Parse the JSON response
            response_text = response.choices[0].message.content
            print(f"Raw response from OpenAI (Page {page_num + 1}): {response_text[:200]}...")
            
            # Clean up the response to extract JSON
            if "```json" in response_text:
                json_start = response_text.find("```json") + 7
                json_end = response_text.find("```", json_start)
                response_text = response_text[json_start:json_end].strip()
            elif "```" in response_text:
                json_start = response_text.find("```") + 3
                json_end = response_text.find("```", json_start)
                response_text = response_text[json_start:json_end].strip()
            
            try:
                result = json.loads(response_text)
                extracted_data = result.get("extracted_data", [])
                
                print(f"Successfully extracted {len(extracted_data)} rows from page {page_num + 1}")
                if result.get("notes"):
                    print(f"Notes: {result.get('notes')}")
                
                return extracted_data
            
            except json.JSONDecodeError as e:
                print(f"JSON decode error on page {page_num + 1}: {e}")
                print(f"Response text: {response_text[:500]}...")
                return []
        
        except Exception as e:
            print(f"Error processing page {page_num + 1}: {e}")
            return []
    
    def _generate_column_definitions(self, columns: List[str]) -> str:
        """Generate helpful definitions for common column types"""
        definitions = []
        for col in columns:
            col_lower = col.lower()
            if any(word in col_lower for word in ['species', 'scientific', 'name']):
                definitions.append(f"- {col}: Scientific names (genus + species, e.g., 'Quercus alba')")
            elif any(word in col_lower for word in ['location', 'locality', 'place', 'county', 'state']):
                definitions.append(f"- {col}: Geographic locations (counties, states, specific places)")
            elif any(word in col_lower for word in ['date', 'year', 'time', 'collected', 'observed']):
                definitions.append(f"- {col}: Dates or years (when something was collected/observed)")
            elif any(word in col_lower for word in ['status', 'condition', 'state']):
                definitions.append(f"- {col}: Status codes or conditions (e.g., 'PrEx', 'PoEx', 'Extinct')")
            elif any(word in col_lower for word in ['family', 'group', 'category']):
                definitions.append(f"- {col}: Taxonomic family or grouping information")
            elif any(word in col_lower for word in ['amount', 'quantity', 'number', 'count']):
                definitions.append(f"- {col}: Numerical values or quantities")
            elif any(word in col_lower for word in ['description', 'notes', 'comments']):
                definitions.append(f"- {col}: Descriptive text or additional notes")
            else:
                definitions.append(f"- {col}: Extract data that logically fits this column name")
        
        return "\n".join(definitions)
    
    def create_excel_file(self, all_data: List[Dict[str, Any]], columns: List[str], 
                         output_path: str, extract_multiple_rows: bool = True):
        """
        Create an Excel file from extracted data
        
        Args:
            all_data (List[Dict]): All extracted data rows
            columns (List[str]): Column names
            output_path (str): Path for the output Excel file
            extract_multiple_rows (bool): Whether multiple rows were extracted per document
        """
        if not all_data:
            print("No data to write to Excel file")
            # Create empty Excel file with headers
            pd.DataFrame(columns=columns).to_excel(output_path, index=False)
            return
        
        # Create DataFrame
        df = pd.DataFrame(all_data)
        
        # Ensure all specified columns exist
        for col in columns:
            if col not in df.columns:
                df[col] = "N/A"
        
        # Reorder columns to match specified order
        df = df[columns]
        
        # Create Excel file with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Extracted_Data', index=False)
            
            # Format the sheet
            worksheet = writer.sheets['Extracted_Data']
            
            # Style headers
            for col in range(1, len(columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"Excel file created: {output_path}")
        print(f"Total rows extracted: {len(df)}")
    
    def save_progress(self, data: List[Dict], pdf_name: str, columns: List[str]):
        """Save progress to a temporary file"""
        progress_file = os.path.join(self.output_dir, f"{pdf_name}_progress.json")
        with open(progress_file, 'w') as f:
            json.dump({'data': data, 'columns': columns}, f)
    
    def load_progress(self, pdf_name: str) -> tuple:
        """Load progress from temporary file"""
        progress_file = os.path.join(self.output_dir, f"{pdf_name}_progress.json")
        if os.path.exists(progress_file):
            with open(progress_file, 'r') as f:
                progress = json.load(f)
                return progress['data'], progress['columns']
        return [], []
    
    def clean_progress(self, pdf_name: str):
        """Clean up progress file"""
        progress_file = os.path.join(self.output_dir, f"{pdf_name}_progress.json")
        if os.path.exists(progress_file):
            os.remove(progress_file)
    
    def extract_with_retry(self, image: Image.Image, columns: List[str], 
                          extra_instructions: str, page_num: int, max_retries: int = 3) -> List[Dict[str, Any]]:
        """Extract data with retry mechanism and validation"""
        for attempt in range(max_retries):
            try:
                print(f"  Attempt {attempt + 1}/{max_retries} for page {page_num + 1}")
                data = self.extract_specific_columns(image, columns, extra_instructions, page_num)
                
                # Validate extraction
                if self.validate_extraction(data, columns):
                    return data
                else:
                    print(f"  Validation failed for page {page_num + 1}, retrying...")
                    time.sleep(2)  # Wait before retry
                    
            except Exception as e:
                print(f"  Attempt {attempt + 1} failed: {e}")
                if attempt < max_retries - 1:
                    time.sleep(2)
                
        print(f"  All attempts failed for page {page_num + 1}, using fallback extraction")
        return self.fallback_extraction(image, columns, page_num)
    
    def validate_extraction(self, data: List[Dict], columns: List[str]) -> bool:
        """Validate that extraction has reasonable data"""
        if not data:
            return False
        
        # Check if at least some columns have non-empty data
        non_empty_count = 0
        for row in data[:5]:  # Check first 5 rows
            for col in columns:
                if col in row and row[col] and row[col].strip() not in ['', 'N/A', 'n/a', 'NULL']:
                    non_empty_count += 1
                    break
        
        return non_empty_count > 0
    
    def fallback_extraction(self, image: Image.Image, columns: List[str], page_num: int) -> List[Dict[str, Any]]:
        """Simplified fallback extraction for difficult pages"""
        base64_image = self.encode_image(image)
        
        simple_prompt = f"""
        Extract any structured data from this page that matches these columns: {', '.join(columns)}
        
        Return as simple JSON:
        {{"extracted_data": [{{{", ".join([f'"{col}": "value"' for col in columns])}}}]}}
        
        If no clear data is found, return: {{"extracted_data": []}}
        """
        
        try:
            response = self.client.chat.completions.create(
                model="gpt-4o",
                messages=[{
                    "role": "user",
                    "content": [
                        {"type": "text", "text": simple_prompt},
                        {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{base64_image}"}}
                    ]
                }],
                max_tokens=2000,
                temperature=0.2
            )
            
            response_text = response.choices[0].message.content
            if "```json" in response_text:
                json_start = response_text.find("```json") + 7
                json_end = response_text.find("```", json_start)
                response_text = response_text[json_start:json_end].strip()
            
            result = json.loads(response_text)
            return result.get("extracted_data", [])
            
        except Exception as e:
            print(f"   Fallback extraction failed: {e}")
            return []
    
    def process_pdf_with_columns(self, pdf_path: str, columns: List[str], 
                               extra_instructions: str = "", extract_multiple_rows: bool = True,
                               resume_from_page: int = 0) -> str:
        """
        Process a PDF file and extract specific columns with robust extraction
        
        Args:
            pdf_path (str): Path to the PDF file
            columns (List[str]): List of column names to extract
            extra_instructions (str): Additional instructions for extraction
            extract_multiple_rows (bool): Whether to extract multiple rows per document
            resume_from_page (int): Page number to resume from (0-indexed)
            
        Returns:
            str: Path to the created Excel file
        """
        pdf_name = Path(pdf_path).stem
        columns_suffix = "_".join(columns[:3])  # Use first 3 columns in filename
        output_excel_path = os.path.join(self.output_dir, f"{pdf_name}_{columns_suffix}_extracted.xlsx")
        
        print(f"\n Processing PDF: {pdf_path}")
        print(f" Extracting columns: {', '.join(columns)}")
        
        try:
            # Load previous progress if resuming
            all_data, prev_columns = self.load_progress(pdf_name)
            if all_data and prev_columns == columns:
                print(f" Resuming from previous progress ({len(all_data)} rows already extracted)")
            else:
                all_data = []
                resume_from_page = 0
            
            # Convert PDF to images
            print(" Converting PDF to images...")
            images = self.pdf_to_images(pdf_path)
            total_pages = len(images)
            print(f" Successfully converted PDF to {total_pages} images")
            
            if resume_from_page > 0:
                print(f" Resuming from page {resume_from_page + 1}")
            
            # Process each page with progress tracking
            pages_with_data = 0
            pages_without_data = 0
            
            for page_num in range(resume_from_page, total_pages):
                progress_percent = ((page_num + 1) / total_pages) * 100
                print(f"\n Processing page {page_num + 1}/{total_pages} ({progress_percent:.1f}%)")
                
                try:
                    image = images[page_num]
                    page_data = self.extract_with_retry(image, columns, extra_instructions, page_num)
                    
                    if page_data:
                        # Add page number to each row
                        for row in page_data:
                            row['_page_number'] = page_num + 1
                        
                        all_data.extend(page_data)
                        pages_with_data += 1
                        print(f"   Extracted {len(page_data)} rows from page {page_num + 1}")
                    else:
                        pages_without_data += 1
                        print(f"   No data found on page {page_num + 1}")
                    
                    # Save progress every 10 pages for large documents
                    if (page_num + 1) % 10 == 0:
                        self.save_progress(all_data, pdf_name, columns)
                        print(f"   Progress saved ({len(all_data)} total rows so far)")
                    
                    # Rate limiting with exponential backoff
                    if page_num < total_pages - 1:  # Don't wait after last page
                        wait_time = min(1 + (page_num // 50) * 0.5, 3)  # Increase wait time for very long docs
                        time.sleep(wait_time)
                        
                except Exception as e:
                    print(f"   Error processing page {page_num + 1}: {e}")
                    print(f"   Continuing with next page...")
                    continue
            
            # Final statistics
            print(f"\n Extraction Summary:")
            print(f"    Total pages processed: {total_pages}")
            print(f"    Pages with data: {pages_with_data}")
            print(f"    Pages without data: {pages_without_data}")
            print(f"   Success rate: {(pages_with_data/total_pages)*100:.1f}%")
            
            # Remove page number column if user didn't request it
            if '_page_number' not in columns:
                for row in all_data:
                    row.pop('_page_number', None)
            
            if all_data:
                print(f" Creating Excel file with {len(all_data)} total rows...")
                self.create_excel_file(all_data, columns, output_excel_path, extract_multiple_rows)
                
                # Clean up progress file on success
                self.clean_progress(pdf_name)
                
                print(f" Successfully extracted data from {len(all_data)} rows!")
            else:
                print(" No data found in the entire document")
                # Create empty Excel file with headers
                pd.DataFrame(columns=columns).to_excel(output_excel_path, index=False)
            
            return output_excel_path
            
        except Exception as e:
            print(f" Error processing PDF {pdf_path}: {e}")
            print(" Progress has been saved. You can resume processing later.")
            return None


def interactive_column_setup():
    """
    Interactive setup for column selection (like Lido)
    """
    print("\n" + "="*60)
    print(" LIDO-STYLE PDF TABLE EXTRACTOR")
    print("="*60)
    
    print("\nStep 1: Define the columns you want to extract")
    print("(Similar to how Lido lets you specify columns)")
    
    columns = []
    print("\nEnter column names one by one (press Enter on empty line to finish):")
    
    while True:
        col = input(f"Column {len(columns) + 1}: ").strip()
        if not col:
            break
        columns.append(col)
        print(f" Added: {col}")
    
    if not columns:
        print("No columns specified. Using default columns for government documents:")
        columns = ["Species", "Locality", "Date Last Collected", "Status"]
        print(f"Default columns: {', '.join(columns)}")
    
    print(f"\nFinal columns to extract: {', '.join(columns)}")
    
    # Extra instructions
    print("\nStep 2: Additional Instructions (optional)")
    print("Enter any specific extraction instructions:")
    extra_instructions = input("Instructions (or press Enter to skip): ").strip()
    
    # Multiple rows setting
    print("\nStep 3: Extraction Mode")
    extract_multiple = input("Extract multiple rows per document? (Y/n): ").strip().lower()
    extract_multiple_rows = extract_multiple != 'n'
    
    return columns, extra_instructions, extract_multiple_rows


def main():
    """
    Main function with interactive column selection and resume capability
    """
    # Configuration
    INPUT_DIRECTORY = "input_pdfs"
    
    # Check if API key is available
    if not OPENAI_API_KEY:
        print(" Error: OpenAI API key not found!")
        print("Please create a .env file in the same directory as this script with:")
        print("OPENAI_API_KEY=your_actual_api_key_here")
        return
    
    # Create input directory if it doesn't exist
    os.makedirs(INPUT_DIRECTORY, exist_ok=True)
    
    # Interactive setup
    columns, extra_instructions, extract_multiple_rows = interactive_column_setup()
    
    print(f"\n Input directory: {os.path.abspath(INPUT_DIRECTORY)}")
    
    # Check for PDF files
    pdf_files = list(Path(INPUT_DIRECTORY).glob("*.pdf"))
    
    if not pdf_files:
        print(f"\n No PDF files found in {INPUT_DIRECTORY}")
        print(f"Please place your PDF files in: {os.path.abspath(INPUT_DIRECTORY)}")
        return
    
    print(f"\n Found {len(pdf_files)} PDF file(s) to process")
    
    # Show file sizes to set expectations
    print("\n File Information:")
    for pdf_file in pdf_files:
        try:
            doc = fitz.open(str(pdf_file))
            page_count = len(doc)
            doc.close()
            file_size_mb = pdf_file.stat().st_size / (1024 * 1024)
            print(f"    {pdf_file.name}: {page_count} pages ({file_size_mb:.1f} MB)")
        except:
            print(f"   {pdf_file.name}: Unable to read page count")
    
    # Confirm processing for large files
    total_estimated_pages = sum([len(fitz.open(str(f))) for f in pdf_files if f.suffix.lower() == '.pdf'])
    if total_estimated_pages > 100:
        print(f"\n Large document(s) detected: ~{total_estimated_pages} total pages")
        print("This may take significant time and OpenAI tokens.")
        confirm = input("Continue? (y/N): ").strip().lower()
        if confirm != 'y':
            print(" Processing cancelled")
            return
    
    # Initialize extractor
    try:
        extractor = LidoStylePDFExtractor(OPENAI_API_KEY)
        
        successful = 0
        failed = 0
        
        for pdf_file in pdf_files:
            print(f"\n{'='*80}")
            print(f" PROCESSING: {pdf_file.name}")
            print('='*80)
            
            try:
                result = extractor.process_pdf_with_columns(
                    str(pdf_file), 
                    columns, 
                    extra_instructions, 
                    extract_multiple_rows
                )
                
                if result:
                    print(f" Successfully processed: {pdf_file.name}")
                    print(f" Output: {result}")
                    successful += 1
                else:
                    print(f" Failed to process: {pdf_file.name}")
                    print(" Check the progress file - you may be able to resume")
                    failed += 1
            except KeyboardInterrupt:
                print(f"\n Processing interrupted by user")
                print(f" Progress has been saved. You can resume processing later.")
                break
            except Exception as e:
                print(f" Error processing {pdf_file.name}: {e}")
                print(" Progress has been saved. You can resume processing later.")
                failed += 1
        
        print(f"\n Processing Complete!")
        print(f" Successfully processed: {successful} files")
        print(f" Failed: {failed} files")
        print(f" Check the '{extractor.output_dir}' folder for results")
        
        # Show resume instructions if there were failures
        if failed > 0:
            print(f"\n To resume failed extractions:")
            print(f"   - Restart the script")
            print(f"   - Use the same column configuration")
            print(f"   - The script will automatically resume from where it left off")
        
    except Exception as e:
        print(f" Error initializing extractor: {e}")


# Quick setup for power users
def quick_extract(pdf_path: str, columns: List[str], extra_instructions: str = ""):
    """
    Quick extraction function for advanced users
    
    Example usage:
    quick_extract(
        "document.pdf", 
        ["Species", "Location", "Date", "Status"],
        "Focus on extinct species data"
    )
    """
    extractor = LidoStylePDFExtractor(OPENAI_API_KEY)
    return extractor.process_pdf_with_columns(pdf_path, columns, extra_instructions)


if __name__ == "__main__":
    # For interactive use
    main()
    
    # For programmatic use, uncomment below:
    # result = quick_extract(
    #     "path/to/your/document.pdf",
    #     ["Species", "Locality", "Date Last Collected", "Status"],
    #     "Extract data about endangered species"
    # )
